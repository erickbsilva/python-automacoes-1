from openpyxl import load_workbook, Workbook

lista_arquivos = ["test", "gastos", "gastos_logo"]

# 1 - Cria nova planilha
wb = Workbook()
nome_arquivo = "files/resultado.xlsx"
for nome in lista_arquivos:
    print(nome)
    arquivo = load_workbook(filename=f"files/{nome}.xlsx")
    sheet = arquivo[nome]
    max_linhas = sheet.max_row
    max_colunas = sheet.max_column

    ws = wb.create_sheet(title=nome)

    # Iterar valores das planilhas
    for i in range(1, max_linhas + 1):
        for j in range(1, max_colunas + 1):
            data = sheet.cell(row=i, column=j)
            ws.cell(row=i, column=j).value = data.value

wb.remove(wb["Sheet"])
wb.save(nome_arquivo)
